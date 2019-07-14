import openpyxl
from openpyxl import load_workbook
from openpyxl import worksheet
from pandas import *

class DataManager:

 # this method is made to teturn the max number of rows
 def getRowCount(path,sheetName):
    Workbook = load_workbook(path, use_iterators=True)
    return Workbook[sheetName].max_row

 # this method is to return the max number of columns
 def getColumnCount(self,path,sheetname):
     Workbook= load_workbook(path,sheetname)
     return Workbook[sheetname].max_column

 # this method is to return the cell data
 def getcellData(path,sheet,col,row):
     Workbook= load_workbook(path,sheet)
     return Workbook[sheet].cell(col,row)

 # this method is to return the execution flag row number
 def get_execution_flag_rownum(path):
     list =[]
     workbook= load_workbook(path)
     sheet=workbook['Driver']
     for row in sheet.iter_rows():
         for cell in row:
             if cell.value =="YES":
                 list.append(cell.row)
     return list


 # this method is to return the Data_Sheet names based on the execution flag
 def get_Data_Sheet_Name(path):
    list=[]
    worksheet=load_workbook(path)['Driver']
    for item in DataManager.get_execution_flag_rownum(path):
        list.append(worksheet.cell(row=item,column=7).value)
    return (list)


 # this method is to return the value of TestDataSheetRowNo from the driver sheet for
 # the (yes) flagged tests
 def get_Data_Sheet_RowNum(path):
    list=[]
    worksheet=load_workbook(path)['Driver']
    for item in DataManager.get_execution_flag_rownum(path):
        list.append(worksheet.cell(row=item,column=8).value)
    return (list)

 def get_Data_Rows(path,index):
     sheet=load_workbook(path)[DataManager.get_Data_Sheet_Name(path)[index]]
     pandasSheet = ExcelFile(path)
     data =pandasSheet.parse(pandasSheet.sheet_names[1])
     dict={}
     print(data.to_dict().get('RowID')[0])


     # for i in range(sheet.columns):
     keys=[]

     for row in sheet.iter_rows(1,1):
      for cell in row:
          keys.append(cell.value)
        # print (cell.value,end=',')
        # print()
     # print(keys)

     # for row in sheet.iter_rows(2,sheet.max_row):
     #     for cell in row:
     #         dict[keys]=cell.value
     #
     # print(dict)


     # dict[sheet.iter_rows(1,1)]

     # keys=[]
     # # for col in range(sheet.max_column):
     # for columns in sheet.iter_cols(1):
     #      for cell in columns:
     #       keys.append(cell.value)
     # print(keys)




     # for row in range(sheet.max_row):
     #     dict[sheet.cell(1,row+1).value]=sheet.cell(row+1,1).value
     #

     #
     # for key,value in dict:
     #     print(key)
     #     print(value)

     # print(sheet.max_column, sheet.max_row)





     #
     # for row in worksheet.iter_rows():
     #     for cell in row:
     #       print (cell.value,end=',')
     #     print()
     #
     #








#print(DataManager.getcellData('/my_Stuff/python/sheet1.xlsx','Sheet1',4,3).value)

# DataManager.get_execution_flag_rownum('/my_Stuff/python/Untitled 1.xlsx')

# print(DataManager.get_execution_flag_rownum('C:/python/DataMgmt.xlsx'))
# print(DataManager.get_Data_Sheet_Name('DataMgmt.xlsx')[0])
# DataManager.get_Data_Sheet_RowNum('DataMgmt.xlsx')

DataManager.get_Data_Rows('DataMgmt.xlsx',1)






